import openpyxl
book = openpyxl.load_workbook("mi_libro.xlsx")
sheet = book.active
sheet2 = book.create_sheet("Inventario")
inventario = [
    ["Producto","Stock"],
    ["Manzanas", 100],
    ["Naranjas", 80]
]
for i in inventario:
    sheet2.append(i)
book.save("mi_libro.xlsx")
print("Archivo modificado")
book = openpyxl.load_workbook("mi_libro.xlsx")
diccionario_ventas = {}
diccionario_inventario = {}
Hventas = book["Ventas"]
for j in Hventas.iter_rows(min_row=2,values_only=True):
    producto,cantidad,precio,total = j
    diccionario_ventas[producto] = {"Cantidad" : cantidad, "Precio" : precio, "Total" : total}
Hinventario = book["Inventario"]
for k in Hinventario.iter_rows(min_row=2,values_only=True):
    prod, stock = k
    diccionario_inventario[prod] = stock
print(diccionario_inventario)
print(diccionario_ventas)